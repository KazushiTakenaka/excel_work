import sys
import os
import shutil
from copy import copy

# Ensure we can import the extraction tool
sys.path.append(os.path.dirname(__file__))
import extract_quotation_details_v3 as extractor

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Please install it.")
    sys.exit(1)

TARGET_FILE = r'c:\Users\taman\OneDrive\デスクトップ\作業中\104.ai_workspaces\excel_work\見積書\加工品リスト.xlsx'

def get_all_data():
    print("Extracting data from files...")
    # Temporarily redirect stdout to suppress extraction logs
    # or just let it print
    
    # We need to manually duplicate the logic of main() in extract_quotation_details_v3 
    # because main() there isn't returning a giant list usable here easily without modifying it.
    # Actually, let's just use the functions directly.
    
    target_dir = extractor.TARGET_DIR
    if not os.path.exists(target_dir):
        print("Target directory not found.")
        return []
        
    all_items = []
    for file in os.listdir(target_dir):
        path = os.path.join(target_dir, file)
        items = []
        if file.lower().endswith(('.xlsx', '.xls')):
            items = extractor.extract_from_excel(path)
        elif file.lower().endswith('.pdf'):
            items = extractor.extract_from_pdf(path)
            
        if items:
            for item in items:
                # Attach source filename for reference
                item['source_file'] = file
                all_items.append(item)
                
    return all_items

def copy_style(source_cell, target_cell):
    """Copy style from source to target cell"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def update_excel(data_items):
    print(f"Updating {os.path.basename(TARGET_FILE)}...")
    
    if not os.path.exists(TARGET_FILE):
        print("Target Excel file not found!")
        return

    wb = openpyxl.load_workbook(TARGET_FILE)
    ws = wb.active
    
    # Find insertion point
    # Look for the first empty cell in Column A (Index 1)
    # OR the row that contains "総合計"
    
    insert_row_idx = -1
    last_data_row_idx = -1
    
    # Iterate rows to find where to append
    # Assuming header is row 1.
    for row in ws.iter_rows(min_row=2):
        cell_a = row[0] # Column A
        val_a = str(cell_a.value).strip() if cell_a.value else ""
        
        # Check for Total row
        # Check all cells in row? User said "上の行を参考にして"
        # Usually total is in a specific column or just somewhere at the bottom
        row_vals = [str(c.value) if c.value else "" for c in row]
        if "総合計" in row_vals:
            insert_row_idx = cell_a.row
            break
        
        # If A is empty, this might be the insertion point
        if not val_a or val_a == 'nan' or val_a == 'None':
            insert_row_idx = cell_a.row
            break
            
        last_data_row_idx = cell_a.row

    if insert_row_idx == -1:
        # Append to the very end
        insert_row_idx = ws.max_row + 1
        
    print(f"Inserting {len(data_items)} items starting at row {insert_row_idx}")
    
    # We will insert new rows. 
    # Note: insert_rows inserts *above* the specified row index.
    ws.insert_rows(insert_row_idx, amount=len(data_items))
    
    # Reference row for coding style (the row above insertion)
    ref_row_idx = insert_row_idx - 1
    if ref_row_idx < 2: ref_row_idx = 2 # Fallback to first data row if existing is empty
    
    # Mapping
    # A: 図面番号 (1)
    # C: 名称 (3)
    # I: 単価① (9)
    # O: 合計 (15)
    # J: 加工先① (10) - Optional, identifying from filename
    
    vendor_map = {
        '26AA0788': '株式会社メイカーズ',
        'QTKG': '創業實業(中国)有限公司',
        'MT05': 'TKエンジニアリング',
        '注文No': 'TKエンジニアリング'
    }
    
    for i, item in enumerate(data_items):
        current_row = insert_row_idx + i
        
        # Determine vendor from filename
        vendor_name = ""
        fname = item.get('source_file', '')
        for key, val in vendor_map.items():
            if key in fname:
                vendor_name = val
                break
        
        # Values to set
        # Column indices are 1-based in openpyxl
        ws.cell(row=current_row, column=1).value = item.get('part_no', '')
        ws.cell(row=current_row, column=3).value = item.get('name', '')
        ws.cell(row=current_row, column=9).value = item.get('unit_price', '')
        ws.cell(row=current_row, column=15).value = item.get('amount', '')
        
        if vendor_name:
             ws.cell(row=current_row, column=10).value = vendor_name

        # Copy style from reference row
        # Iterate all columns in the row
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=ref_row_idx, column=col)
            target_cell = ws.cell(row=current_row, column=col)
            copy_style(source_cell, target_cell)

    wb.save(TARGET_FILE)
    print("Update complete.")

def main():
    data = get_all_data()
    print(f"DEBUG: Found {len(data)} items total.")
    if not data:
        print("No data found to update.")
        return

    update_excel(data)

if __name__ == "__main__":
    main()
